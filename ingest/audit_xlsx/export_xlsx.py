"""파싱 결과 행들을 XLSX 파일로 작성."""
from __future__ import annotations

from pathlib import Path
from typing import Any, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


_FONT_NAME = "Arial"
_HEADER_FILL = PatternFill("solid", start_color="305496")
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFF")
_BODY_FONT = Font(name=_FONT_NAME)

# (key, header, width, wrap)
_COLUMNS: tuple[tuple[str, str, int, bool], ...] = (
    ("stock_code", "종목코드", 10, False),
    ("corp_code", "DART고유번호", 12, False),
    ("corp_name", "회사명", 22, False),
    ("market", "시장구분", 14, False),
    ("report_kind", "보고서 종류", 16, False),
    ("attach_title", "첨부 제목", 26, False),
    ("parent_rcept_no", "감사보고서제출 접수번호", 20, False),
    ("biz_report_rcept_no", "사업보고서 접수번호", 20, False),
    ("stlm_dt", "결산기준일", 12, False),
    ("bsns_year", "사업연도", 10, False),
    ("adt_opinion", "감사의견", 14, False),
    ("adtor", "감사인(회계법인)", 22, False),
    ("cpa_partner_name", "업무수행 공인회계사", 18, False),
    ("core_adt_matter", "핵심감사사항(KAM)", 60, True),
    ("emphs_matter", "강조사항", 50, True),
    ("other_matters", "기타사항", 50, True),
    ("audit_report_body", "감사보고서 본문 전체", 80, True),
    ("raw_text_length", "원문 길이(태그포함)", 14, False),
    ("flat_text_length", "평문 길이", 12, False),
    ("audit_report_body_length", "본문 슬라이스 길이", 14, False),
    ("skip_reason", "스킵 사유", 22, False),
    ("parse_error", "파싱 오류", 30, True),
)


def _write_summary_sheet(wb: Workbook, rows: Sequence[dict[str, Any]]) -> None:
    sheet = wb.active
    sheet.title = "summary"
    headers = [c[1] for c in _COLUMNS]
    sheet.append(headers)
    for col_idx, (_key, _header, width, wrap) in enumerate(_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        sheet.column_dimensions[letter].width = width
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, (key, _header, _w, wrap) in enumerate(_COLUMNS, start=1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=row.get(key))
            cell.font = _BODY_FONT
            cell.alignment = Alignment(
                vertical="top", wrap_text=wrap, horizontal="left"
            )

    sheet.freeze_panes = "A2"


def write_xlsx(rows: Sequence[dict[str, Any]], out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    _write_summary_sheet(wb, rows)
    wb.save(out_path)
    return out_path
